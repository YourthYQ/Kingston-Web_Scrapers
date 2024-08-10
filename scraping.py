import sys
sys.setrecursionlimit(sys.getrecursionlimit() * 5)

import tkinter as tk
from tkinter import messagebox
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

# Functions to fetch data
def get_gigabyte_consumer_data():
    import Gigabyte_Consumer as Gigabyte_Consumer
    return Gigabyte_Consumer.ProductHandler([
        "https://www.gigabyte.com/Motherboard/All-Series", 
        "https://www.gigabyte.com/Gaming-PC", 
        "https://www.gigabyte.com/Mini-PcBarebone/All", 
        "https://www.gigabyte.com/Laptop/All-Series"
    ], {}, {}).get_combined_data()

def get_gigabyte_enterprise_data():
    import Gigabyte_Enterprise as Gigabyte_Enterprise
    return Gigabyte_Enterprise.ProductHandler([
        'https://www.gigabyte.com/Enterprise/Server-Motherboard',
        'https://www.gigabyte.com/Enterprise/Workstation-Motherboard',
        'https://www.gigabyte.com/Enterprise/Rack-Server',
        'https://www.gigabyte.com/Enterprise/GPU-Server',
        'https://www.gigabyte.com/Enterprise/High-Density-Server',
        'https://www.gigabyte.com/Enterprise/Tower-Server'
    ], {}, {}).get_data_from_response()

def get_asus_data():
    import Asus_Motherboard as Asus_Motherboard
    return Asus_Motherboard.main()

def get_asrock_data():
    import ASRock as ASRock
    return ASRock.ProductHandler([
        'https://www.asrock.com/mb/index.us.asp#AllProduct',
        'https://www.asrock.com/nettop/index.us.asp#AllProduct'
    ]).get_data_from_response()

def get_asrock_rack_data():
    import ASRock_Rack as ASRock_Rack
    return ASRock_Rack.ProductHandler([
        'https://www.asrockrack.com/general/products-ajax.asp?Model=&Type=Server&CPU=&Category=&Usage=&Socket=&CPUnum=&DIMM=&Form=&BuildingBlocks=&Bricks=&Form2=&Acc=&Life=&p=1',
        'https://www.asrockrack.com/general/products-ajax.asp?Model=&Type=WS&CPU=&Category=&Usage=&Socket=&CPUnum=&DIMM=&Form=&BuildingBlocks=&Bricks=&Form2=&Acc=&Life=&p=1'
    ]).get_data_from_response()

# Function to combine data into a single Excel file
def combine_excels(selected_sources, combined_path):
    with pd.ExcelWriter(combined_path, engine='openpyxl') as writer:
        data_written = False
        if 'Gigabyte Consumer' in selected_sources:
            gigabyte_consumer_df = get_gigabyte_consumer_data()
            if not gigabyte_consumer_df.empty:
                gigabyte_consumer_df.to_excel(writer, sheet_name='Gigabyte_Consumer', index=False)
                data_written = True
        
        if 'Gigabyte Enterprise' in selected_sources:
            gigabyte_enterprise_df = get_gigabyte_enterprise_data()
            if not gigabyte_enterprise_df.empty:
                gigabyte_enterprise_df.to_excel(writer, sheet_name='Gigabyte_Enterprise', index=False)
                data_written = True
        
        if 'ASUS' in selected_sources:
            asus_df = get_asus_data()
            if not asus_df.empty:
                asus_df.to_excel(writer, sheet_name='ASUS', index=False)
                data_written = True

        if 'ASRock' in selected_sources:
            asrock_df = get_asrock_data()
            if not asrock_df.empty:
                asrock_df.to_excel(writer, sheet_name='ASRock', index=False)
                data_written = True

        if 'ASRock Rack' in selected_sources:
            asrock_rack_df = get_asrock_rack_data()
            if not asrock_rack_df.empty:
                asrock_rack_df.to_excel(writer, sheet_name='ASRock Rack', index=False)
                data_written = True
    
    if data_written:
        format_excel(combined_path)
        # Displays a message box indicating the success of the export
        messagebox.showinfo("Success", f"Data has been exported to {combined_path}")
    else:
        messagebox.showwarning("Warning", "No data to export. Please check your sources.")

# Function to format the Excel file
def format_excel(path):
    # Loads the workbook from the specified path
    wb = load_workbook(path)
    
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
    
    wb.save(path)

# Function to start data scraping based on selected sources
def start_scraping():
    selected_sources = [source for source, var in source_vars.items() if var.get() == 1]
    if not selected_sources:
        messagebox.showwarning("Warning", "Please select at least one source.")
        return
    combine_excels(selected_sources, "Combined_Products.xlsx")

# Initialize the GUI application
app = tk.Tk()
app.title("Product Scraper")
app.geometry("500x400")  # Set the window size to 500x400 pixels
app.config(bg="lightgrey")

# Set the window icon
app.iconbitmap("app_icon.ico")

# Create a title label
title_label = tk.Label(app, text="Product Data Scraper", font=("Helvetica", 22, "bold"), bg="lightgrey", fg="black")
title_label.grid(row=0, column=0, columnspan=3, pady=10)

# Create an instruction label
instruction_label = tk.Label(app, text="Select the sources you want to scrape data from:", font=("Helvetica", 14), bg="lightgrey", fg="black")
instruction_label.grid(row=1, column=0, columnspan=3, pady=5)

# Define sources
sources = ['Gigabyte Consumer', 'Gigabyte Enterprise', 'ASUS', 'ASRock', 'ASRock Rack']
source_vars = {source: tk.IntVar() for source in sources}

# Create a frame for checkbuttons with scrollbar
frame = tk.Frame(app, bg="lightgrey")
frame.grid(row=2, column=0, columnspan=3, padx=10, pady=10)

canvas = tk.Canvas(frame, bg="lightgrey", width=300)
scrollbar = tk.Scrollbar(frame, orient="vertical", command=canvas.yview)
scrollable_frame = tk.Frame(canvas, bg="lightgrey")

scrollable_frame.bind(
    "<Configure>",
    lambda e: canvas.configure(
        scrollregion=canvas.bbox("all")
    )
)

canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
canvas.configure(yscrollcommand=scrollbar.set)

for i, source in enumerate(sources):
    chk = tk.Checkbutton(scrollable_frame, text=source, variable=source_vars[source], bg="lightgrey", fg="black", font=("Helvetica", 16))
    chk.pack(anchor="w", padx=10, pady=2)

canvas.pack(side="left", fill="both", expand=True)
scrollbar.pack(side="right", fill="y")

# Create a button to start scraping with border-radius effect
start_button = tk.Button(app, text="Start Scraping", command=start_scraping, bg="white", fg="black", font=("Helvetica", 16), borderwidth=1, relief=tk.SOLID, padx=10, pady=5)
start_button.grid(row=3, column=0, columnspan=3, pady=20)

# Create a status label
status_label = tk.Label(app, text="", font=("Helvetica", 16), bg="lightgrey", fg="black")
status_label.grid(row=4, column=0, columnspan=3, pady=5)

# Configure grid weights for centering
app.grid_columnconfigure(0, weight=1)
app.grid_columnconfigure(1, weight=1)
app.grid_columnconfigure(2, weight=1)
app.grid_rowconfigure(0, weight=1)
app.grid_rowconfigure(1, weight=1)
app.grid_rowconfigure(2, weight=1)
app.grid_rowconfigure(3, weight=1)
app.grid_rowconfigure(4, weight=1)

# Start the GUI main loop
app.mainloop()